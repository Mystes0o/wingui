import openpyxl
import time
import loguru


def excel_to_list(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    # print( data)
    return data

def excel_to_list2(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []

    rows = sheet.iter_rows(min_row=2, values_only=True)
    for row in rows:
        filtered_row = [cell for cell in row if cell is not None]
        data.append(filtered_row)
    # print(data)
    return data

def list_to_excel(file_path, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    start_row = 2
    for row_index, (cpu_mem_gpu, video_params, video_file_path) in enumerate(data):
        row_num = start_row + row_index

        sheet[f'G{row_num}'] = cpu_mem_gpu

        formatted_video_info = "\n".join(video_params)
        sheet[f'H{row_num}'] = formatted_video_info
        sheet[f'I{row_num}'] = video_file_path

    now_time = time.strftime("%Y-%m-%d_%H%M%S", time.localtime())
    save_path = file_path.replace('模板', now_time)
    loguru.logger.info(f"录制数据结果：{save_path}")
    workbook.save(save_path)


if __name__ == '__main__':
    # excel_to_list2(r"C:\Users\admini\Desktop\ERE性能测试用例_模板.xlsx")

    test_list = [['平均CPU使用率1.87%\n平均内存使用率211.73M\n平均GPU使用率3.87%', ['format:mov,mp4,m4a,3gp,3g2,mj2', 'duration:9.920000', 'bit_rate:772', 'width:3840', 'height:2560', 'fps:25.0', 'codec:h264', 'pixel_format:yuv420p', 'frame_count:51', 'rotate:None'], 'E:\\python\\wingui\\case\\2025_07_11_164628\\20250711_164642.mp4'], ['平均CPU使用率24.83%\n平均内存使用率257.87M\n平均GPU使用率1.82%', ['format:mov,mp4,m4a,3gp,3g2,mj2', 'duration:10.333334', 'bit_rate:931', 'width:3840', 'height:2560', 'fps:30.0', 'codec:h264', 'pixel_format:yuv420p', 'frame_count:217', 'rotate:None'], 'E:\\python\\wingui\\case\\2025_07_11_164628\\20250711_164709.mp4']]
    list_to_excel(r"E:\python\wingui\case\ERE性能测试用例_模板.xlsx", test_list)